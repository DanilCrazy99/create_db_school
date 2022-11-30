# -*- coding: utf-8 -*-

import openpyxl
import xlrd

# имя загружаемого файла
tmp_path = 'timetable.xls'


def create_timetable_list_xlsx(path=tmp_path):
    """
    Чтение только файла XLSX
    :param path: Путь к файлу
    :return:
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    # получить имена страниц с расписаниями. Название листа содержит сочетание "-параллель"
    all_name_sheet = wb.get_sheet_names()
    print('all_sheet= ', all_name_sheet)
    str_find = '-параллель'
    # максимальное кол-во уроков в день
    count_lesson_day = 15
    list_class = []
    list_lessons = [[]]
    date_time = '01-01-2020'
    for item in all_name_sheet:
        if item.find(str_find) > 0:
            sheet = wb.get_sheet_by_name(item)
            # sheet = wb.active
            # print(sheet[1][0].value)   A1    ABCDEFGHI  INDEX[ЦИФРА][БУКВА]
            # print(sheet[3][4].value)   E3    123456789
            # info_editor = sheet[2][1].value   Информация: Ответственный и дата
            lesson_cell_coordinate_number = 3
            date_time = sheet[2][5].value
            for columns_name_class in range(3, sheet.max_column):  # Получаю список классов при помощи прохода по 3-й строке
                if sheet[3][columns_name_class].value:
                    list_class.append(sheet[3][columns_name_class].value)

            for columns in range(3, sheet.max_column):  # переборка по координате "буква"
                for rows in range(3, (7*count_lesson_day)):  # переборка по координате "цифра"

                    if sheet[rows][columns].value:  # Если выбранная ячейка содержит данные
                        if sheet[rows][columns].value in list_class:  # Значение ячейки есть в списке с классами
                            name_class = sheet[rows][columns].value

                            for i in range(count_lesson_day):  # Записываем count_lesson_day строк под ним
                                incr = len(list_lessons) - 1
                                if sheet[rows + i + 1][columns].value is None:  # Убираю значения None из списков
                                    pass
                                else:
                                    lesson_cell_coordinate_number = rows + i + 1
                                    if sheet[lesson_cell_coordinate_number][1].value == 'урок':
                                        break  # прерываем цикл при натыкании на столбец 'урок'

                                    list_lessons[incr].append(name_class)
                                    list_lessons[incr].append(str(sheet[lesson_cell_coordinate_number][1].value))
                                    list_lessons[incr].append(sheet[lesson_cell_coordinate_number][columns].value)
                                    list_lessons[incr].append(sheet[rows+1][0].value)
                                list_lessons.append([])
    return list_lessons, date_time

def create_timetable_list(path=tmp_path):
    """
    Чтение только файла XLS
    :param path: Путь к файлу
    :return:
    """
    wb = xlrd.open_workbook_xls(path, formatting_info=True)
    # получить имена страниц с расписаниями. Название листа содержит сочетание "-параллель"
    all_name_sheet = wb.sheet_names()
    str_find = '-параллель'
    # максимальное кол-во уроков в день
    count_lesson_day = 15
    list_class = []
    list_lessons = [[]]
    date_time = '01-01-2020'

    for item in all_name_sheet:
        if item.find(str_find) > 0:  # перебираем только листы в названии которых есть слово "-параллель"
            sheet = wb.sheet_by_name(item)  # получаем данные с листа
            max_row = sheet.nrows  # получаем сколько строк занято данными
            max_column = sheet.ncols  # получаем число столбцов на листе
            # sheet = wb.active
            # print(sheet[1][0].value)   A1    ABCDEFGHI  INDEX[ЦИФРА][БУКВА]
            # print(sheet[3][4].value)   E3    123456789
            # info_editor = sheet[2][1].value   Информация: Ответственный и дата
            lesson_cell_coordinate_number = 3
            date_time = xlrd.xldate_as_datetime(xldate=sheet[1][10].value, datemode=0).date().isoformat()  # получаем дату листа

            for columns_name_class in range(4, max_column, 3):  # Создаем список классов на листе.
                if sheet[2][columns_name_class].value:   # проверка существования данных в ячейке
                    list_class.append(sheet[2][columns_name_class].value)  # добавляем описание класса "цифра буква"

            for columns in range(4, max_column, 3):  # переборка по столбцам листа "буква"
                for rows in range(2, 7*count_lesson_day):  # переборка по строкам "номер урока" всех дней недели

                    if max_row <= rows:
                        continue

                    if sheet[rows][columns].value:  # Если выбранная ячейка содержит данные
                        if sheet[rows][columns].value in list_class:  # Значение ячейки есть в списке с классами
                            name_class = sheet[rows][columns].value

                            for i in range(count_lesson_day):  # Записываем count_lesson_day строк под ним
                                incr = len(list_lessons) - 1
                                # прерываем цикл по достижении конца таблицы с данными
                                if max_row <= (rows + i + 1):
                                    continue

                                if sheet[rows + i + 1][columns].value is None \
                                        or sheet[rows + i + 1][columns].value == '':  # Убираю значения None из списков
                                    continue
                                else:
                                    lesson_cell_coordinate_number = rows + i + 1
                                    if sheet[lesson_cell_coordinate_number][1].value == 'урок':
                                        break  # прерываем цикл при натыкании на столбец 'урок'

                                    list_lessons[incr].append(name_class)  # описание класса
                                    list_lessons[incr].append(int(sheet[lesson_cell_coordinate_number][1].value))  # номер урока
                                    if sheet.cell_type(lesson_cell_coordinate_number, columns+1) == 2:
                                        number_cab = '(каб.' \
                                                     + str(int(sheet[lesson_cell_coordinate_number][columns + 1].value)) \
                                                     + ')'
                                    else:
                                        number_cab = sheet[lesson_cell_coordinate_number][columns+1].value
                                        if len(number_cab) > 0:
                                            number_cab = '(' + number_cab + ')'
                                    tmp_lesson = sheet[lesson_cell_coordinate_number][3].value \
                                                 + ' ' + sheet[lesson_cell_coordinate_number][columns].value \
                                                 + f' {number_cab}'
                                    list_lessons[incr].append(tmp_lesson)  # описание урока
                                    list_lessons[incr].append(sheet[rows+1][0].value)  # день недели урока
                                list_lessons.append([])

    return list_lessons, date_time


def create_postgres_list(list_func):
    completed_list = []
    for a in range(len(list_func)):
        if len(list_func[a]) != 0:
            completed_list.append(list_func[a])
    return completed_list

if __name__ == '__main__':
    create_timetable_list()
