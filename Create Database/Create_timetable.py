import openpyxl
from Variables.Var_database import path_timetable

wb = openpyxl.load_workbook(path_timetable)
sheet = wb.active
# print(sheet[1][0].value)   A1    ABCDEFGHI  INDEX[ЦИФРА][БУКВА]
# print(sheet[3][4].value)   E3    123456789
list_class = []
list_lessons = []

for columns_name_class in range(3, sheet.max_column):  # Получаю список классов при помощи прохода по 3-й строке
    if sheet[3][columns_name_class].value:
        list_class.append(sheet[3][columns_name_class].value)

for columns in range(3, sheet.max_column):  # переборка по координате "буква"
    for rows in range(3, 52):  # переборка по координате "цифра"
        day_week_val = int(rows/8)
        day_week = sheet[4 + day_week_val * 8][0].value
        if sheet[rows][columns].value:  # Если выбранная ячейка содержит данные
            if sheet[rows][columns].value in list_class:  # Значение ячейки есть в списке с классами
                name_class = sheet[rows][columns].value

                for i in range(7):  # Записываем 7 строк под ним
                    if sheet[rows + i + 1][columns].value is None:  # Убираю значения None из списков
                        list_lessons.append(f'{name_class}Y_')
                    else:
                        lesson_cell_coordinate_number = rows + i + 1
                        list_lessons.append(str(name_class) +
                                            'Y' +
                                            str(sheet[lesson_cell_coordinate_number][1].value)[0] +
                                            'Y' +
                                            str(sheet[rows + i + 1][columns].value) +
                                            'Y' +
                                            str(day_week)
                                            )
                    if i == 6:
                        list_lessons.append(f'{name_class}Y_')

print(list_class)
print(list_lessons)


def create_postgres_list():
    completed_list = [
        [1, 2, 3, 4, 5, 6]]  # класс, урок(время), предмет, день недели, время(актуализации), ответственный
    return completed_list  # Должно вернуть список который можно будет вставить в функцию INSERT

# for i in list_class['class']:
#     if i == '5Б':
#         print(list_class['class'].index(i))
